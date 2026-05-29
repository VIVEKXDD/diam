"""
map_supplier.py
---------------
Takes a raw supplier Excel file as input and outputs a new Excel file
with all column names renamed to canonical field names.

The mapping rules now live in a single flat JSON (mapping_flat.json) that
maps each canonical field name to ALL possible source aliases across every
supplier.  You no longer need to know which supplier produced a file — the
script auto-detects the right rename by checking which aliases actually exist
in the input.

Usage
-----
    python map_supplier.py <input_file>... [--supplier <name>] [--output <path>] [--master <path>]

Arguments
---------
    input_file          Path to one or more raw supplier Excel files (.xlsx)
    --supplier <name>   Optional. One of: Ratnakala | Vaibhav | Glowstar |
                        Zhaveri | Karigar.  Repeat once per input file or once
                        to apply the same supplier hint to all inputs.
    --output <path>     Optional. Output .xlsx path. Defaults to a generated
                        path when omitted.
    --master <path>     Optional. Existing master workbook to append
                        normalized rows to.

Examples
--------
    # Auto-detect supplier from column names (no --supplier needed)
    python map_supplier.py zhaveri_raw.xlsx

    # Provide supplier explicitly for richer warnings
    python map_supplier.py zhaveri_raw.xlsx --supplier Zhaveri

    # Specify output path too
    python map_supplier.py data/vaibhav.xlsx --supplier Vaibhav --output data/out.xlsx

    # Append multiple normalized uploads to an existing master workbook
    python map_supplier.py vaibhav.xlsx zhaveri.xlsx --master master.xlsx

Notes
-----
- Columns that match no canonical alias are DROPPED from the output.
- A 'supplier' column is prepended to every output sheet (value = supplier
  name when given, else "UNKNOWN").
- A lightweight data normalization layer cleans values after mapping:
  removes common unit suffixes (e.g. "5 mm" → "5"), trims whitespace,
  converts numeric strings to numbers, and imputes missing / zero values.
  Price fields are exempt from median imputation and are left blank when
  missing.
- The rules file (mapping_flat.json) must be in the same directory as this
  script.
- If the input file has multiple sheets, all sheets are processed.
- Collision handling: if the same source column name is an alias for more
  than one canonical field (e.g. "Amount" appears under both price_per_carat
  and price), the script warns you and picks the first canonical field whose
  alias list contains it. Pass --supplier to get the correct resolution.
"""

# ── Standard library ──────────────────────────────────────────────────────────
import sys
import json
import argparse
import re
from pathlib import Path

# ── Third-party ───────────────────────────────────────────────────────────────
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Rules file location ───────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
RULES_FILE = SCRIPT_DIR / "mapping_flat.json"

# Supplier-aware priority: when two canonical fields share a source alias,
# we prefer the one that is correct for the given supplier.
# Format: { supplier_lower: { source_alias_lower: preferred_canonical } }
SUPPLIER_ALIAS_PRIORITY: dict[str, dict[str, str]] = {
    "vaibhav":   {"amount": "price_per_carat"},   # Vaibhav "Amount" = per-carat price
    "ratnakala": {"amount": "price"},              # Ratnakala "Amount" = total price
    "glowstar":  {"amount": "price"},
    "zhaveri":   {"amount": "price", "depth": "depth_pct"},
    "karigar":   {},
}


# =============================================================================
# load_rules
# =============================================================================
def load_rules(rules_path: Path) -> tuple[dict[str, list[str]], dict]:
    """
    Read mapping_flat.json and return:
      c2a  — canonical_to_all_suppliers  { canonical: [alias1, alias2, …] }
      warn — warnings                    { canonical: { supplier: text } }
    """
    with open(rules_path, "r", encoding="utf-8") as fh:
        data = json.load(fh)

    c2a  = data.get("canonical_to_all_suppliers", {})
    warn = data.get("warnings", {})

    if not isinstance(c2a, dict):
        raise ValueError("mapping_flat.json: 'canonical_to_all_suppliers' must be a dict.")

    return c2a, warn


# =============================================================================
# build_alias_map
# =============================================================================
def normalize_alias(alias: str) -> str:
    return str(alias).strip().lower()


def build_alias_map(
    c2a: dict[str, list[str]],
    supplier: str | None,
) -> dict[str, str]:
    """
    Invert c2a into  { normalized_source_alias: canonical_field }.

    Collision handling
    ------------------
    When the same alias maps to more than one canonical field
    (e.g. "Amount" → price_per_carat  AND  price), we resolve via:
      1. If supplier is known and SUPPLIER_ALIAS_PRIORITY has an entry → use it.
      2. Otherwise keep the first canonical field encountered (dict order)
         and print a WARNING so the user knows.

    Returns a dict keyed by normalized alias strings.
    """
    alias_map: dict[str, str] = {}           # normalized alias → canonical
    collision_log: dict[str, list[str]] = {} # normalized alias → [canonical, canonical, …]

    sup_lower = supplier.lower() if supplier else ""
    priority  = SUPPLIER_ALIAS_PRIORITY.get(sup_lower, {})

    for canonical, aliases in c2a.items():
        for alias in aliases:
            normalized = normalize_alias(alias)
            if normalized in alias_map:
                collision_log.setdefault(normalized, [alias_map[normalized]])
                collision_log[normalized].append(canonical)

                if normalized in priority and priority[normalized] == canonical:
                    alias_map[normalized] = canonical
            else:
                alias_map[normalized] = canonical

    for alias, candidates in collision_log.items():
        chosen = alias_map[alias]
        others = [c for c in candidates if c != chosen]
        print(
            f"  [WARNING] COLLISION: source column '{alias}' matches multiple canonical "
            f"fields {candidates}.\n"
            f"     Resolved to '{chosen}'. Others ignored: {others}.\n"
            f"     Pass --supplier to get supplier-aware resolution."
        )

    return alias_map


# =============================================================================
# map_df
# =============================================================================
def map_df(
    df: pd.DataFrame,
    alias_map: dict[str, str],
    warn: dict,
    supplier: str | None,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    """
    Rename columns in *df* using *alias_map* and collect audit lists.

    Returns
    -------
    mapped_df    — renamed DataFrame with "supplier" column prepended
    mapped_cols  — ["source  →  canonical", …] for the report sheet
    dropped_cols — column names that had no alias match
    """
    sup_label = supplier.upper() if supplier else "UNKNOWN"
    sup_lower = supplier.lower() if supplier else ""

    mapped_cols:  list[str] = []
    dropped_cols: list[str] = []
    rename_map:   dict[str, str] = {}

    for col in df.columns:
        normalized_col = normalize_alias(col)
        if normalized_col in alias_map:
            canonical = alias_map[normalized_col]
            rename_map[col] = canonical
            mapped_cols.append(f"{col}  →  {canonical}")

            field_warnings = warn.get(canonical, {})
            if sup_lower and sup_lower.capitalize() in field_warnings:
                msg = field_warnings[sup_lower.capitalize()]
                print(f"  [WARNING] '{col}' → '{canonical}': {msg}")
            elif not sup_lower and field_warnings:
                for s, msg in field_warnings.items():
                    print(f"  [WARNING] '{col}' → '{canonical}' [{s}]: {msg}")
        else:
            dropped_cols.append(col)

    out = df.drop(columns=dropped_cols).rename(columns=rename_map)
    out = out.loc[:, ~out.columns.duplicated()]
    out.insert(0, "supplier", sup_label)
    out = clean_df(out)

    return out, mapped_cols, dropped_cols


# =============================================================================
# data cleaning
# =============================================================================

def normalize_value(value):
    if pd.isna(value):
        return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None

        text = text.replace(",", "")
        text = re.sub(r"(?i)\b(mm|cts?|carat|carats|percent|pct|%)\b", "", text)
        text = text.strip()

        if not text:
            return None

        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
            num = float(text)
            return int(num) if num.is_integer() else num

        return text

    if isinstance(value, (np.integer, np.floating, int, float)):
        if np.isnan(value):
            return None
        if value == 0:
            return 0
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    return value


def find_similar_rows(df: pd.DataFrame, row_idx: int, threshold: float = 0.5) -> list[int]:
    """
    Find rows similar to the given row based on matching non-null values.
    Returns list of similar row indices sorted by similarity score (descending).
    """
    target_row = df.iloc[row_idx]
    similarities = []

    for idx, other_row in df.iterrows():
        if idx == row_idx:
            continue

        matches = 0
        comparisons = 0

        for col in df.columns:
            target_val = target_row[col]
            other_val = other_row[col]

            if pd.isna(target_val) or pd.isna(other_val):
                continue

            comparisons += 1
            if str(target_val).lower().strip() == str(other_val).lower().strip():
                matches += 1

        if comparisons > 0:
            score = matches / comparisons
            if score >= threshold:
                similarities.append((idx, score))

    similarities.sort(key=lambda x: x[1], reverse=True)
    return [idx for idx, _ in similarities]


def impute_strings_by_similarity(df: pd.DataFrame) -> pd.DataFrame:
    """
    Impute missing string values by finding similar rows and using their values.
    """
    imputed = df.copy()
    str_cols = imputed.select_dtypes(include=["string", "object"]).columns

    for row_idx in range(len(imputed)):
        for col in str_cols:
            val = imputed.iloc[row_idx][col]

            if pd.isna(val) or (isinstance(val, str) and val.strip() == ""):
                similar_rows = find_similar_rows(imputed, row_idx, threshold=0.5)

                for similar_idx in similar_rows:
                    similar_val = imputed.iloc[similar_idx][col]
                    if pd.notna(similar_val) and (not isinstance(similar_val, str) or similar_val.strip() != ""):
                        imputed.iloc[row_idx, imputed.columns.get_loc(col)] = similar_val
                        break

    return imputed


def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.copy()
    cleaned = cleaned.apply(lambda col: col.map(normalize_value))

    for col in cleaned.columns:
        values = cleaned[col].dropna()
        if not values.empty and all(isinstance(v, (int, float, np.integer, np.floating)) for v in values):
            cleaned[col] = pd.to_numeric(cleaned[col], errors="coerce")

    for col in cleaned.select_dtypes(include=[np.number]).columns:
        if "price" in str(col).lower():
            continue
        column = cleaned[col].replace(0, np.nan)
        non_zero = column.dropna()
        if not non_zero.empty:
            fill_value = non_zero.median()
            cleaned[col] = column.fillna(fill_value)

    cleaned = impute_strings_by_similarity(cleaned)

    for col in cleaned.select_dtypes(include=["string", "object"]).columns:
        cleaned[col] = cleaned[col].where(
            cleaned[col].notna() & cleaned[col].astype(str).str.strip().ne(""),
            "UNKNOWN"
        )

    return cleaned


# =============================================================================
# write_excel  (unchanged from original)
# =============================================================================
def write_excel(output_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    import numpy as np

    wb = Workbook()
    wb.remove(wb.active)

    thin   = Side(style="thin",   color="BBBBBB")
    medium = Side(style="medium", color="444444")

    def bdr():     return Border(top=thin,   bottom=thin,   left=thin,   right=thin)
    def hdr_bdr(): return Border(top=medium, bottom=medium, left=medium, right=medium)

    header_fill   = PatternFill("solid", start_color="1F4E79")
    alt_fill      = PatternFill("solid", start_color="EAF2F8")
    supplier_fill = PatternFill("solid", start_color="D5F5E3")

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font     = Font(name="Arial", size=9)
    supplier_font = Font(name="Arial", size=9, bold=True, color="145A32")

    for sheet_name, df in sheets.items():
        ws   = wb.create_sheet(title=sheet_name[:31])
        cols = list(df.columns)

        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font      = header_font
            c.fill      = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border    = hdr_bdr()
        ws.row_dimensions[1].height = 28

        for ri, (_, row) in enumerate(df.iterrows()):
            fill = alt_fill if ri % 2 == 1 else PatternFill()
            for ci, col in enumerate(cols, 1):
                val = row[col]
                if pd.isna(val):
                    val = None
                elif isinstance(val, float) and not pd.isna(val) and val == int(val):
                    val = int(val)

                c = ws.cell(row=ri + 2, column=ci, value=val)
                if col == "supplier":
                    c.font  = supplier_font
                    c.fill  = supplier_fill
                else:
                    c.font  = data_font
                    c.fill  = fill
                c.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=False)
                c.border = bdr()
            ws.row_dimensions[ri + 2].height = 15

        for ci, col in enumerate(cols, 1):
            values  = [str(col)] + [str(v) for v in df[col].dropna()]
            max_len = max((len(v) for v in values), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

        ws.freeze_panes = "B2"

    wb.save(output_path)


# =============================================================================
# append_report_sheet  (updated: warnings are now keyed by canonical field)
# =============================================================================
def append_report_sheet(
    output_path: Path,
    suppliers: list[str] | None,
    mapped_cols: list[str],
    dropped_cols: list[str],
    warn: dict,
) -> None:
    if suppliers and len(set(suppliers)) == 1:
        sup_label = suppliers[0].upper()
        sup_cap   = suppliers[0].capitalize()
    else:
        sup_label = "MULTIPLE" if suppliers else "UNKNOWN"
        sup_cap = None

    wb = load_workbook(output_path)
    if "_mapping_report" in wb.sheetnames:
        del wb["_mapping_report"]
    ws = wb.create_sheet(title="_mapping_report")

    def bdr():
        s = Side(style="thin", color="BBBBBB")
        return Border(top=s, bottom=s, left=s, right=s)

    # Title
    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value=f"Mapping Report — {sup_label}")
    t.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    t.fill      = PatternFill("solid", start_color="1C2833")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    row = 3

    # ── Section 1: Renamed ────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="RENAMED COLUMNS").font = Font(
        name="Arial", bold=True, size=10)
    row += 1

    for ci, h in enumerate(["Source Column", "Canonical Field", "Warning"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()
    row += 1

    alt = PatternFill("solid", start_color="EAF2F8")

    for i, entry in enumerate(mapped_cols):
        source, canonical = [x.strip() for x in entry.split("→")]
        # Look up warning for this canonical field + current supplier
        field_warn = warn.get(canonical, {})
        warning    = field_warn.get(sup_cap, "") if sup_cap else \
                     " | ".join(f"[{s}] {m}" for s, m in field_warn.items())

        fill = alt if i % 2 == 1 else PatternFill()
        for ci, val in enumerate([source, canonical, warning], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = Font(name="Arial", size=9,
                               color="C0392B" if warning and ci == 3 else "000000")
            c.fill      = fill
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = bdr()
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1

    # ── Section 2: Dropped ────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="DROPPED COLUMNS (no mapping)").font = Font(
        name="Arial", bold=True, size=10)
    row += 1

    for ci, h in enumerate(["Column Name", "Reason"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", start_color="7B241C")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()
    row += 1

    drop_fill = PatternFill("solid", start_color="FADBD8")
    for col in dropped_cols:
        c1 = ws.cell(row=row, column=1, value=col)
        c2 = ws.cell(row=row, column=2, value="Not in mapping rules — dropped")
        for c in [c1, c2]:
            c.font      = Font(name="Arial", size=9)
            c.fill      = drop_fill
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = bdr()
        ws.row_dimensions[row].height = 15
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 70

    wb.save(output_path)


# =============================================================================
# main
# =============================================================================
def make_sheet_title(base_name: str, existing: set[str]) -> str:
    candidate = base_name[:31]
    counter = 1
    while candidate in existing:
        suffix = f"_{counter}"
        candidate = base_name[:31-len(suffix)] + suffix
        counter += 1
    return candidate


def order_columns(df: pd.DataFrame, c2a: dict[str, list[str]]) -> pd.DataFrame:
    desired = ["supplier"] + [canonical for canonical in c2a.keys() if canonical in df.columns]
    extras = [col for col in df.columns if col not in desired]
    return df.loc[:, desired + extras]


def load_master_sheet(master_path: Path, sheet_name: str = "MASTER") -> pd.DataFrame | None:
    if not master_path.exists():
        return None
    xl = pd.ExcelFile(master_path)
    if sheet_name in xl.sheet_names:
        return pd.read_excel(master_path, sheet_name=sheet_name)
    return None


def main() -> None:
    # ── CLI ───────────────────────────────────────────────────────────────────
    parser = argparse.ArgumentParser(
        description="Map supplier Excel columns to canonical field names.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("input_files", nargs="+", help="One or more raw supplier .xlsx files")
    parser.add_argument(
        "--output", "-o",
        help="Output .xlsx path. If omitted, defaults to a generated path or master path when using --master.",
    )
    parser.add_argument(
        "--supplier", "-s",
        action="append",
        default=[],
        help="Supplier name (Ratnakala|Vaibhav|Glowstar|Zhaveri|Karigar). "
             "Repeat once per input file or once to apply to all files.",
    )
    parser.add_argument(
        "--master", "-m",
        help="Path to master workbook to append normalized rows to. If omitted, a new mapped workbook is created.",
    )
    args = parser.parse_args()

    input_paths = [Path(path) for path in args.input_files]
    supplier_args = [s.strip() for s in args.supplier if s.strip()]
    if supplier_args and len(supplier_args) not in (1, len(input_paths)):
        print("ERROR: --supplier must be provided once or once per input file.")
        raise SystemExit(1)

    supplier_list: list[str | None]
    if len(supplier_args) == 1:
        supplier_list = [supplier_args[0]] * len(input_paths)
    elif len(supplier_args) == len(input_paths):
        supplier_list = supplier_args
    else:
        supplier_list = [None] * len(input_paths)

    master_path = Path(args.master) if args.master else None

    if args.master:
        output_path = Path(args.output) if args.output else Path(args.master)
    else:
        output_path = Path(args.output) if args.output else (
            input_paths[0].parent / ("combined_mapped.xlsx" if len(input_paths) > 1 else f"{input_paths[0].stem}_mapped.xlsx")
        )

    # ── Validate ──────────────────────────────────────────────────────────────
    for input_path in input_paths:
        if not input_path.exists():
            print(f"ERROR: Input file not found: {input_path}")
            raise SystemExit(1)
    if not RULES_FILE.exists():
        print(f"ERROR: Rules file not found: {RULES_FILE}")
        raise SystemExit(1)

    # ── Load rules ────────────────────────────────────────────────────────────
    print(f"\nLoading mapping rules from: {RULES_FILE.name}")
    c2a, warn = load_rules(RULES_FILE)

    print(f"Input files   : {', '.join(str(p) for p in input_paths)}")
    if master_path:
        print(f"Master file   : {master_path}")
    print(f"Output file   : {output_path}")
    print(f"Canonical fields in rules : {len(c2a)}")

    alias_map = build_alias_map(c2a, None)

    all_sheets  : dict[str, pd.DataFrame] = {}
    all_mapped  : list[str] = []
    all_dropped : list[str] = []

    for input_path, supplier in zip(input_paths, supplier_list):
        print(f"\nProcessing file: {input_path}")
        if supplier:
            print(f"  Supplier hint: {supplier}")
        xl = pd.ExcelFile(input_path)

        for sheet_name in xl.sheet_names:
            print(f"\n── Sheet: '{sheet_name}' ──")
            raw_df = pd.read_excel(input_path, sheet_name=sheet_name)
            print(f"   {len(raw_df)} rows × {len(raw_df.columns)} columns")

            alias_map = build_alias_map(c2a, supplier)
            mapped_df, mapped_cols, dropped_cols = map_df(
                raw_df, alias_map, warn, supplier
            )

            print(f"   Mapped  : {len(mapped_cols)} columns")
            print(f"   Dropped : {len(dropped_cols)} columns "
                  f"({', '.join(dropped_cols) if dropped_cols else 'none'})")

            sheet_title = sheet_name if len(input_paths) == 1 else make_sheet_title(
                f"{input_path.stem}_{sheet_name}", set(all_sheets.keys())
            )
            all_sheets[sheet_title] = mapped_df
            all_mapped += mapped_cols
            all_dropped += [f"{input_path.stem}:{sheet_name}: {c}" for c in dropped_cols]

    if master_path:
        existing_master = load_master_sheet(master_path)
        master_tables = []
        if existing_master is not None:
            master_tables.append(existing_master)
        master_tables.extend(all_sheets.values())

        if master_tables:
            master_df = pd.concat(master_tables, ignore_index=True, sort=False)
            master_df = order_columns(master_df, c2a)
        else:
            master_df = pd.DataFrame(columns=["supplier"] + list(c2a.keys()))

        output_sheets = {"MASTER": master_df}
    else:
        output_sheets = all_sheets

    print(f"\nWriting output...")
    write_excel(output_path, output_sheets)
    append_report_sheet(output_path, [s for s in supplier_list if s], all_mapped, all_dropped, warn)

    print(f"\n✓ Done.  Output saved to: {output_path}")
    print(f"  Sheets: {list(output_sheets.keys())} + _mapping_report")


if __name__ == "__main__":
    main()